import pandas as pd, numpy as np, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import os
HERE=os.path.dirname(os.path.abspath(__file__)); ROOT=os.path.dirname(HERE)
O=os.path.join(ROOT,'state')+'/'
OUT=os.path.dirname(ROOT)+'/'
df=pd.read_csv(O+'combined3.csv')
df['posted_date']=pd.to_datetime(df['posted_date'],errors='coerce')

ARIAL=Font(name='Arial',size=10); HDR=Font(name='Arial',size=10,bold=True,color='FFFFFF')
HFILL=PatternFill('solid',fgColor='1F3864'); TITLE=Font(name='Arial',size=14,bold=True,color='1F3864')
H2=Font(name='Arial',size=11,bold=True,color='1F3864'); SUB=Font(name='Arial',size=9,italic=True,color='595959')
thin=Side(style='thin',color='D9D9D9'); BOX=Border(left=thin,right=thin,top=thin,bottom=thin)
LINK=Font(name='Arial',size=9,color='0563C1',underline='single')
WARN=PatternFill('solid',fgColor='FFF2CC')

COLS=[('source','Source',11),('model_full','Make & Model',22),('variant','Variant',26),
      ('year','Year',7),('price_lacs','Price (lacs)',12),('price_pkr','Price (PKR)',13),
      ('mileage_km','Mileage (km)',13),('transmission','Transmission',13),('city','City',14),
      ('area','Area / Locality',30),('posted_date','Posted / Updated',17),('ago','Ad age',9),
      ('freshness','Freshness',18),('badge','Seller Badge',22),('rating','Inspection /10',14),
      ('pics','Photos',8),('body','Body Type',16),('flags','Notes / Flags',40),('url','Listing URL',66)]

def sheet(ws,frame,name,start=1,link_col='url'):
    keys=[c[0] for c in COLS]
    for j,(k,h,w) in enumerate(COLS,start=1):
        c=ws.cell(start,j,h); c.font=HDR; c.fill=HFILL; c.border=BOX
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width=w
    for i,(_,r) in enumerate(frame.iterrows(),start=start+1):
        for j,(k,h,w) in enumerate(COLS,start=1):
            v=r.get(k)
            if pd.isna(v): v=None
            if k=='posted_date' and v is not None: v=pd.Timestamp(v).to_pydatetime()
            c=ws.cell(i,j,v); c.font=ARIAL; c.border=BOX
            if k=='price_lacs': c.number_format='0.00'
            elif k=='price_pkr': c.number_format='#,##0'
            elif k=='mileage_km': c.number_format='#,##0'
            elif k=='year': c.number_format='0'; c.alignment=Alignment(horizontal='center')
            elif k=='posted_date': c.number_format='dd-mmm-yyyy'
            elif k=='rating': c.number_format='0.0'
            elif k in ('pics','ago','freshness'): c.alignment=Alignment(horizontal='center')
            if k=='url' and v: c.hyperlink=v; c.font=LINK
            if r.get('flags'): 
                if k=='flags': c.fill=WARN
    ws.freeze_panes=ws.cell(start+1,4)
    if name:
        ref=f'A{start}:{get_column_letter(len(COLS))}{len(frame)+start}'
        t=Table(displayName=name,ref=ref); t.tableStyleInfo=TableStyleInfo(name='TableStyleLight1',showRowStripes=True)
        ws.add_table(t)

wb=Workbook()

# ---- Read Me ----
ws=wb.active; ws.title='Read Me'
rows=[('660cc Japanese Kei Cars - Islamabad & Rawalpindi',None),
('Auto-refreshed every 3 hours, 9am-9pm, while the Claude app is open. Sources: PakWheels and OLX Pakistan. Companion file: kei_cars_islamabad_rawalpindi.html (browsable gallery with photos).',None),(None,None),
('Search criteria',None),
('Engine capacity','600-660 cc'),('Cities','Islamabad, Rawalpindi (plus a small number of nearby towns, flagged)'),
('Max price','PKR 3,000,000 (30 lacs)'),('Model year','2010 and newer'),('Mileage','no limit'),
('Body types','all - hatchback, tall-boy/MPV, van'),(None,None),
('Listing counts',None),
('PakWheels','641 listings. Site has a real 600-660cc filter, so these are reliable. All 26 result pages scraped.'),
('OLX Pakistan','271 listings. OLX has no engine-size filter, so ~40 kei model names were searched per city and then filtered.'),
('Total','912 listings'),(None,None),
('Tabs',None),
('All Listings','Everything, newest ad first. Full clickable URL in the last column.'),
('Fresh This Week','Only ads posted or bumped in the last 7 days - the ones actually still available.'),
('Best Value','Newer + lower mileage + priced at or below the model average.'),
('By Model','Count, cheapest / average / dearest per model. Live formulas.'),
('Price vs Mileage','Median asking price by model and mileage band.'),(None,None),
('Column notes',None),
('Posted / Updated','Calculated back from the relative age the site showed ("3 days ago"), so it is accurate to the day, not the hour. On PakWheels this is when the ad was last bumped, not first posted.'),
('Ad age','Raw relative age from the site. m=minutes, h=hours, d=days, w=weeks, mo=months.'),
('Seller Badge','PakWheels only. "Managed by PakWheels" means PakWheels handles the sale and the car has been inspected.'),
('Inspection /10','PakWheels inspection score where one exists. Blank means the car was never inspected.'),
('Photos','Number of pictures in the ad. Ads with 1-3 photos are worth extra scepticism.'),
('Price (PKR)','Exact figure from the listing. PakWheels rounds to "lacs" on screen; this is the underlying number.'),(None,None),
('Caveats - please read',None),
('1','IMPORTANT FIX vs the first version: OLX quietly ignores the city filter when a keyword has few local matches, so my first pass wrongly included cars in Karachi, Lahore, Sialkot and elsewhere. Those 436 rows have been removed. The OLX count dropped from 730 to 271 as a result. This version is location-verified against the area shown on each ad.'),
('2','31 OLX listings are in Wah, Taxila, Attock, Murree, Jhelum or Gujar Khan rather than Isb/Rwp proper. They are kept but flagged in Notes.'),
('3','OLX ads are free text, so Variant and Mileage are only as good as what the seller typed. Some sellers put two years in the title ("2019 2025" = built 2019, registered 2025); the Year column takes the structured field, not the title.'),
('4','Pakistani-assembled Suzuki Wagon R is 1000cc and was excluded from OLX unless the ad said Stingray / FX / FZ / Hybrid. Some genuine JDM 660cc Wagon Rs may therefore be missing from the OLX rows.'),
('5','Seven listings PakWheels tagged as 600-660cc were actually Mehran, Bolan, Ravi, Carry, Cuore or Kix and have been removed. Model names and variants are now taken from the PakWheels image slug, which is more accurate than the ad title. Flagged rows - implausible mileage, a price too low for the year (usually instalment or leasing ads dressed as sale prices), or a duplicate. Verify before calling.'),
('6','Prices are asking prices. Kei imports in Pakistan typically negotiate 3-8% off ask.'),
('7','Always get an auction sheet verification for a JDM import before paying anything.'),
]
for i,(a,b) in enumerate(rows,start=1):
    ws.cell(i,1,a).font=ARIAL; ws.cell(i,2,b).font=ARIAL
    for j in (1,2): ws.cell(i,j).alignment=Alignment(vertical='top',wrap_text=True)
ws['A1'].font=TITLE; ws['A2'].font=SUB
for r in (4,12,17,24,33): ws.cell(r,1).font=H2
ws.column_dimensions['A'].width=20; ws.column_dimensions['B'].width=105

# ---- All Listings ----
ws2=wb.create_sheet('All Listings'); sheet(ws2,df,'AllListings')

# ---- Fresh This Week ----
fresh=df[df['days']<=7].sort_values('price_lacs')
ws3=wb.create_sheet('Fresh This Week')
ws3['A1']=f'{len(fresh)} ads posted or bumped in the 7 days to 30 July 2026, cheapest first'
ws3['A1'].font=H2
ws3['A2']='Older ads are often already sold. Start here.'; ws3['A2'].font=SUB
sheet(ws3,fresh,'FreshWeek',start=4)

# ---- Best Value ----
avg=df.groupby('model_full')['price_lacs'].transform('mean')
bv=df[(df['flags'].fillna('').str.contains('too cheap|too low|implausibly')==False)
      &(df['price_lacs']<=avg)&(df['year']>=2014)
      &(df['mileage_km'].fillna(1e9)<=120000)&(df['mileage_km'].fillna(0)>=5000)].copy()
bv['value_score']=((bv['year']-2009)*2.0-bv['price_lacs']*0.55-bv['mileage_km']/40000).round(2)
bv=bv.sort_values('value_score',ascending=False).head(60)
ws4=wb.create_sheet('Best Value')
ws4['A1']='Top 60 by value score - 2014+, under 120,000 km, priced at or below the model average'; ws4['A1'].font=H2
ws4['A2']='Value Score = (Year-2009)x2 - Price(lacs)x0.55 - Mileage/40,000. Heuristic ranking, not an appraisal.'; ws4['A2'].font=SUB
sheet(ws4,bv,'BestValue',start=4)
vc=len(COLS)+1
c=ws4.cell(4,vc,'Value Score'); c.font=HDR; c.fill=HFILL; c.border=BOX; c.alignment=Alignment(horizontal='center',wrap_text=True)
for i,(_,r) in enumerate(bv.iterrows(),start=5):
    cc=ws4.cell(i,vc,float(r['value_score'])); cc.font=ARIAL; cc.border=BOX; cc.number_format='0.00'
ws4.column_dimensions[get_column_letter(vc)].width=12

# ---- By Model ----
ws5=wb.create_sheet('By Model')
n=len(df)+1
hdrs=['Make & Model','Body Type','Listings','Cheapest (lacs)','Average (lacs)','Dearest (lacs)','Avg mileage (km)','Newest year','Posted in last 7 days']
for j,h in enumerate(hdrs,start=1):
    c=ws5.cell(1,j,h); c.font=HDR; c.fill=HFILL; c.border=BOX; c.alignment=Alignment(horizontal='center',wrap_text=True)
models=sorted(df['model_full'].unique())
fresh_counts=df[df['days']<=7].groupby('model_full').size()
for i,m in enumerate(models,start=2):
    ws5.cell(i,1,m).font=ARIAL
    ws5.cell(i,2,df.loc[df['model_full']==m,'body'].iloc[0]).font=ARIAL
    ws5.cell(i,3,f"=COUNTIF('All Listings'!$B$2:$B${n},$A{i})")
    ws5.cell(i,4,f"=IFERROR(_xlfn.MINIFS('All Listings'!$E$2:$E${n},'All Listings'!$B$2:$B${n},$A{i}),\"\")")
    ws5.cell(i,5,f"=IFERROR(AVERAGEIFS('All Listings'!$E$2:$E${n},'All Listings'!$B$2:$B${n},$A{i}),\"\")")
    ws5.cell(i,6,f"=IFERROR(_xlfn.MAXIFS('All Listings'!$E$2:$E${n},'All Listings'!$B$2:$B${n},$A{i}),\"\")")
    ws5.cell(i,7,f"=IFERROR(AVERAGEIFS('All Listings'!$G$2:$G${n},'All Listings'!$B$2:$B${n},$A{i}),\"\")")
    ws5.cell(i,8,f"=IFERROR(_xlfn.MAXIFS('All Listings'!$D$2:$D${n},'All Listings'!$B$2:$B${n},$A{i}),\"\")")
    ws5.cell(i,9,int(fresh_counts.get(m,0)))
    for j in range(1,10):
        c=ws5.cell(i,j); c.font=ARIAL; c.border=BOX
        if j in (3,7,9): c.number_format='#,##0'
        elif j==8: c.number_format='0'
        elif j in (4,5,6): c.number_format='0.00'
last=len(models)+1
ws5.cell(last+2,1,'Total listings').font=Font(name='Arial',size=10,bold=True)
tc=ws5.cell(last+2,3,f'=SUM(C2:C{last})'); tc.font=Font(name='Arial',size=10,bold=True); tc.number_format='#,##0'
ws5.cell(last+3,1,'Columns C-H are live formulas over the All Listings tab and update if you edit that data.').font=SUB
for j,w in enumerate([24,17,10,15,14,15,18,12,20],start=1): ws5.column_dimensions[get_column_letter(j)].width=w
ws5.freeze_panes='A2'

# ---- Price vs Mileage ----
bands=[(0,50000,'< 50k'),(50000,100000,'50-100k'),(100000,150000,'100-150k'),(150000,10**9,'> 150k')]
def band(v):
    if pd.isna(v): return 'unknown'
    for lo,hi,l in bands:
        if lo<=v<hi: return l
    return 'unknown'
d2=df.copy(); d2['band']=d2['mileage_km'].apply(band)
piv=d2.pivot_table(index='model_full',columns='band',values='price_lacs',aggfunc='median').round(2)
piv=piv.reindex(columns=[b[2] for b in bands]+['unknown'])
piv=piv.dropna(axis=1,how='all')
ws6=wb.create_sheet('Price vs Mileage')
ws6['A1']='Median asking price (lacs) by model and mileage band'; ws6['A1'].font=H2
ws6['A2']='Blank = no listings in that band.'; ws6['A2'].font=SUB
c=ws6.cell(4,1,'Make & Model'); c.font=HDR; c.fill=HFILL; c.border=BOX
for j,cn in enumerate(piv.columns,start=2):
    c=ws6.cell(4,j,cn); c.font=HDR; c.fill=HFILL; c.border=BOX; c.alignment=Alignment(horizontal='center')
for i,(m,row) in enumerate(piv.iterrows(),start=5):
    ws6.cell(i,1,m).font=ARIAL; ws6.cell(i,1).border=BOX
    for j,cn in enumerate(piv.columns,start=2):
        v=row[cn]; cc=ws6.cell(i,j,None if pd.isna(v) else float(v))
        cc.font=ARIAL; cc.border=BOX; cc.number_format='0.00'; cc.alignment=Alignment(horizontal='center')
ws6.column_dimensions['A'].width=24
for j in range(2,len(piv.columns)+2): ws6.column_dimensions[get_column_letter(j)].width=13
ws6.freeze_panes='B5'

wb.save(OUT+'660cc_kei_cars_isb_rwp.xlsx')
print('saved rows',len(df),'fresh',len(fresh),'models',len(models))
